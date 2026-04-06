from fastapi import APIRouter, HTTPException, Depends, Request, UploadFile, File
from fastapi.responses import FileResponse
from pathlib import Path
import shutil
import re
import uuid
import math
import csv
import io
import json
import logging
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from app.core.database import db, next_sequence
from app.models.schemas import *
from app.core.security import hash_password, verify_password, create_access_token
from app.api.dependencies import get_current_user, login_limiter, verify_token
from app.core.utils import parse_date_flexible, safe_str, check_customer_duplicate, check_vehicle_duplicate, create_activity

router = APIRouter()
logger = logging.getLogger(__name__)


# Backup Service Class
class BackupService:
    def __init__(self, db, backup_config: BackupConfig):
        self.db = db
        self.config = backup_config
        self.backup_root = Path(backup_config.backup_location)
        self.backup_root.mkdir(parents=True, exist_ok=True)
    
    async def create_backup(self, user_id: str, backup_type: str = "manual", export_format: str = "json") -> BackupJob:
        """Create a new backup job with specified format"""
        job = BackupJob(
            status="running",
            start_time=datetime.utcnow(),
            created_by=user_id,
            backup_type=backup_type
        )
        
        try:
            # Create backup directory
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_dir = self.backup_root / f"backup_{timestamp}"
            backup_dir.mkdir(parents=True, exist_ok=True)
            
            # Backup all collections
            collections = ['users', 'customers', 'vehicles', 'sales', 'services', 'spare_parts', 'spare_part_bills']
            total_records = 0
            records_by_collection = {}
            collection_data = {}
            
            for collection_name in collections:
                collection = getattr(self.db, collection_name)
                documents = await collection.find().to_list(length=None)
                
                # Convert ObjectId to string for JSON serialization
                for doc in documents:
                    if '_id' in doc:
                        doc['_id'] = str(doc['_id'])
                
                collection_data[collection_name] = documents
                record_count = len(documents)
                records_by_collection[collection_name] = record_count
                total_records += record_count
            
            # Create files based on export format
            if export_format.lower() == "excel":
                await self.create_excel_backup(backup_dir, collection_data, records_by_collection, user_id, backup_type)
            else:
                # Default JSON/CSV format
                await self.create_json_csv_backup(backup_dir, collection_data, records_by_collection, user_id, backup_type)
            
            # Compress backup if enabled
            final_path = str(backup_dir)
            backup_size = 0
            
            if self.config.compress_backups:
                if export_format.lower() == "excel":
                    zip_path = f"{backup_dir}_excel.zip"
                else:
                    zip_path = f"{backup_dir}.zip"
                
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for file_path in backup_dir.rglob('*'):
                        if file_path.is_file():
                            zipf.write(file_path, file_path.relative_to(backup_dir))
                
                # Remove original directory and get compressed size
                shutil.rmtree(backup_dir)
                backup_size = os.path.getsize(zip_path) / (1024 * 1024)  # MB
                final_path = zip_path
            else:
                # Calculate directory size
                backup_size = sum(f.stat().st_size for f in backup_dir.rglob('*') if f.is_file()) / (1024 * 1024)
            
            # Update job with completion details
            job.status = "completed"
            job.end_time = datetime.utcnow()
            job.total_records = total_records
            job.backup_size_mb = round(backup_size, 2)
            job.backup_file_path = final_path
            job.records_backed_up = records_by_collection
            
        except Exception as e:
            job.status = "failed"
            job.end_time = datetime.utcnow()
            job.error_message = str(e)
        
        # Save job to database
        await self.db.backup_jobs.insert_one(job.dict())
        
        return job
    
    async def create_excel_backup(self, backup_dir: Path, collection_data: dict, records_by_collection: dict, user_id: str, backup_type: str):
        """Create Excel backup with multiple sheets"""
        try:
            # Create comprehensive Excel file
            excel_file = backup_dir / "backup_data.xlsx"
            
            # Create workbook and remove default sheet
            workbook = Workbook()
            workbook.remove(workbook.active)
            
            # Define header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Create sheets for each collection
            for collection_name, documents in collection_data.items():
                if not documents:
                    continue
                    
                # Create sheet
                sheet = workbook.create_sheet(title=collection_name.replace('_', ' ').title())
                
                try:
                    # Flatten nested data for Excel compatibility
                    flattened_documents = []
                    for doc in documents:
                        flat_doc = self.flatten_document(doc)
                        flattened_documents.append(flat_doc)
                    
                    # Convert to DataFrame for easier Excel manipulation
                    df = pd.DataFrame(flattened_documents)
                    
                    # Add data to sheet
                    for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                        for col_num, value in enumerate(row_data, 1):
                            # Convert non-serializable values to strings
                            if value is None:
                                value = ""
                            elif isinstance(value, (dict, list)):
                                value = str(value)
                            
                            cell = sheet.cell(row=row_num, column=col_num, value=str(value))
                            
                            # Apply header styling
                            if row_num == 1:
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = header_alignment
                    
                    # Auto-adjust column widths
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        sheet.column_dimensions[column_letter].width = adjusted_width
                        
                except Exception as e:
                    logger.warning(f"Failed to create Excel sheet for {collection_name}: {e}")
                    # Create a simple sheet with error message
                    sheet.cell(row=1, column=1, value=f"Error processing {collection_name}")
                    sheet.cell(row=2, column=1, value=str(e))
            
            # Create summary sheet
            summary_sheet = workbook.create_sheet(title="Backup Summary", index=0)
            
            # Summary data with IST timezone
            current_utc = datetime.utcnow()
            ist_time = current_utc + timedelta(hours=5, minutes=30)
            
            summary_data = [
                ["Backup Information", ""],
                ["Backup Date", ist_time.strftime("%Y-%m-%d %H:%M:%S IST")],
                ["Backup Type", backup_type.title()],
                ["Created By", user_id],
                ["Total Records", sum(records_by_collection.values())],
                ["Export Format", "Excel Workbook"],
                ["Timezone", "IST (UTC+5:30)"],
                ["", ""],
                ["Collection Statistics", "Records"],
            ]
            
            # Add collection statistics
            for collection, count in records_by_collection.items():
                summary_data.append([collection.replace('_', ' ').title(), count])
            
            # Add summary data to sheet
            for row_num, (key, value) in enumerate(summary_data, 1):
                summary_sheet.cell(row=row_num, column=1, value=str(key))
                summary_sheet.cell(row=row_num, column=2, value=str(value))
                
                # Style headers
                if key in ["Backup Information", "Collection Statistics"]:
                    summary_sheet.cell(row=row_num, column=1).font = Font(bold=True, size=14)
                    summary_sheet.cell(row=row_num, column=2).font = Font(bold=True, size=14)
            
            # Auto-adjust summary sheet columns
            summary_sheet.column_dimensions['A'].width = 25
            summary_sheet.column_dimensions['B'].width = 15
            
            # Save workbook
            workbook.save(excel_file)
            logger.info(f"Excel backup created successfully: {excel_file}")
            
        except Exception as e:
            # Fallback to JSON if Excel creation fails
            logger.error(f"Excel backup creation failed: {e}, falling back to JSON")
            await self.create_json_csv_backup(backup_dir, collection_data, records_by_collection, user_id, backup_type)
    
    def flatten_document(self, doc: dict, prefix: str = "") -> dict:
        """Flatten nested dictionary for Excel compatibility"""
        flattened = {}
        
        for key, value in doc.items():
            new_key = f"{prefix}_{key}" if prefix else key
            
            if isinstance(value, dict):
                # Recursively flatten nested dictionaries
                nested_flattened = self.flatten_document(value, new_key)
                flattened.update(nested_flattened)
            elif isinstance(value, list):
                # Convert lists to comma-separated strings
                if value and isinstance(value[0], dict):
                    # For list of dictionaries, create a summary
                    flattened[new_key] = f"[{len(value)} items]"
                    # Add first item details if available
                    if value:
                        first_item = self.flatten_document(value[0], f"{new_key}_item1")
                        flattened.update(first_item)
                else:
                    # For simple lists, join as string
                    flattened[new_key] = ", ".join(str(item) for item in value)
            else:
                # Simple values
                flattened[new_key] = value
        
        return flattened
    
    async def create_json_csv_backup(self, backup_dir: Path, collection_data: dict, records_by_collection: dict, user_id: str, backup_type: str):
        """Create JSON and CSV backup files"""
        for collection_name, documents in collection_data.items():
            # Save as JSON
            json_file = backup_dir / f"{collection_name}.json"
            async with aiofiles.open(json_file, 'w') as f:
                await f.write(json.dumps(documents, default=str, indent=2))
            
            # Save as CSV if data exists
            if documents:
                try:
                    df = pd.DataFrame(documents)
                    csv_file = backup_dir / f"{collection_name}.csv"
                    df.to_csv(csv_file, index=False)
                except Exception as e:
                    logger.warning(f"Failed to create CSV for {collection_name}: {e}")
        
        # Create backup summary with IST timezone
        current_utc = datetime.utcnow()
        ist_time = current_utc + timedelta(hours=5, minutes=30)
        
        summary = {
            'backup_date': ist_time.isoformat(),
            'backup_date_utc': current_utc.isoformat(),
            'timezone': 'IST (UTC+5:30)',
            'total_records': sum(records_by_collection.values()),
            'records_by_collection': records_by_collection,
            'backup_type': backup_type,
            'created_by': user_id
        }
        
        summary_file = backup_dir / 'backup_summary.json'
        async with aiofiles.open(summary_file, 'w') as f:
            await f.write(json.dumps(summary, indent=2))
    
    async def get_backup_stats(self) -> BackupStats:
        """Get backup statistics"""
        jobs = await self.db.backup_jobs.find().to_list(length=None)
        
        total_backups = len(jobs)
        successful_backups = len([j for j in jobs if j['status'] == 'completed'])
        failed_backups = len([j for j in jobs if j['status'] == 'failed'])
        
        last_backup = None
        oldest_backup = None
        total_size = 0
        
        if jobs:
            sorted_jobs = sorted(jobs, key=lambda x: x['created_at'], reverse=True)
            last_backup = sorted_jobs[0]['created_at']
            oldest_backup = sorted_jobs[-1]['created_at']
            
            # Calculate total storage used
            for job in jobs:
                if job['status'] == 'completed':
                    total_size += job.get('backup_size_mb', 0)
        
        return BackupStats(
            total_backups=total_backups,
            successful_backups=successful_backups,
            failed_backups=failed_backups,
            last_backup_date=last_backup,
            oldest_backup_date=oldest_backup,
            total_storage_used_mb=round(total_size, 2)
        )
    
    async def cleanup_old_backups(self, retention_days: int):
        """Clean up backups older than retention period"""
        cutoff_date = datetime.utcnow() - timedelta(days=retention_days)
        
        # Remove old backup files
        for backup_path in self.backup_root.glob('backup_*'):
            if backup_path.stat().st_mtime < cutoff_date.timestamp():
                try:
                    if backup_path.is_file():
                        backup_path.unlink()
                    else:
                        shutil.rmtree(backup_path)
                    
                    # Update database to mark as cleaned up
                    await self.db.backup_jobs.update_one(
                        {"backup_file_path": str(backup_path)},
                        {"$set": {"cleaned_up": True}}
                    )
                except Exception as e:
                    logger.error(f"Failed to cleanup backup {backup_path}: {e}")

# Initialize backup service
backup_service = None

async def get_backup_service():
    """Get or create backup service instance"""
    global backup_service
    
    if backup_service is None:
        # Get backup config from database or create default
        config_doc = await db.backup_config.find_one()
        if not config_doc:
            # Create default config
            default_config = BackupConfig()
            await db.backup_config.insert_one(default_config.dict())
            config = default_config
        else:
            config = BackupConfig(**config_doc)
        
        backup_service = BackupService(db, config)
    
    return backup_service

# Backup API Endpoints
@router.get("/api/backup/config", response_model=BackupConfig)
async def get_backup_config(current_user: dict = Depends(verify_token)):
    """Get backup configuration"""
    config_doc = await db.backup_config.find_one()
    if not config_doc:
        # Create default config
        default_config = BackupConfig()
        await db.backup_config.insert_one(default_config.dict())
        return default_config
    return BackupConfig(**config_doc)

@router.put("/api/backup/config", response_model=BackupConfig)
async def update_backup_config(
    config_update: dict,
    current_user: dict = Depends(verify_token)
):
    """Update backup configuration"""
    config_update['updated_at'] = datetime.utcnow()
    
    result = await db.backup_config.update_one(
        {}, 
        {"$set": config_update},
        upsert=True
    )
    
    # Refresh backup service with new config
    global backup_service
    backup_service = None
    
    updated_config = await db.backup_config.find_one()
    return BackupConfig(**updated_config)

@router.post("/api/backup/create", response_model=BackupJob)
async def create_manual_backup(
    backup_create: BackupJobCreate,
    current_user: dict = Depends(verify_token)
):
    """Create a manual backup"""
    service = await get_backup_service()
    job = await service.create_backup(
        current_user['user_id'], 
        backup_create.backup_type,
        backup_create.export_format
    )
    return job

@router.get("/api/backup/jobs", response_model=List[BackupJob])
async def get_backup_jobs(
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(verify_token)
):
    """Get backup job history"""
    jobs = await db.backup_jobs.find().skip(skip).limit(limit).sort("created_at", -1).to_list(length=None)
    return [BackupJob(**job) for job in jobs]

@router.get("/api/backup/stats", response_model=BackupStats)
async def get_backup_statistics(current_user: dict = Depends(verify_token)):
    """Get backup system statistics"""
    service = await get_backup_service()
    return await service.get_backup_stats()

@router.delete("/api/backup/cleanup")
async def cleanup_old_backups(
    retention_days: int = 30,
    current_user: dict = Depends(verify_token)
):
    """Clean up old backups"""
    service = await get_backup_service()
    await service.cleanup_old_backups(retention_days)
    return {"message": f"Cleanup completed for backups older than {retention_days} days"}

@router.get("/api/backup/download/{job_id}")
async def download_backup(
    job_id: str,
    current_user: dict = Depends(verify_token)
):
    """Download a backup file"""
    from fastapi.responses import FileResponse
    
    job = await db.backup_jobs.find_one({"id": job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Backup job not found")
    
    if job['status'] != 'completed':
        raise HTTPException(status_code=400, detail="Backup is not completed")
    
    backup_path = Path(job['backup_file_path'])
    if not backup_path.exists():
        raise HTTPException(status_code=404, detail="Backup file not found")
    
    return FileResponse(
        path=backup_path,
        filename=backup_path.name,
        media_type='application/octet-stream'
    )
